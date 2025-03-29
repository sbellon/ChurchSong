import abc
import datetime
import enum
import pathlib
import sys
import typing
from collections import defaultdict

import alive_progress  # pyright: ignore[reportMissingTypeStubs]
import openpyxl
import openpyxl.styles
import openpyxl.utils
import prettytable

from churchsong.churchtools import ChurchToolsAPI
from churchsong.configuration import Configuration


# The values of FormatType are the accepted formats of prettytable and openpyxl.
class FormatType(str, enum.Enum):
    TEXT = 'text'
    HTML = 'html'
    JSON = 'json'
    CSV = 'csv'
    LATEX = 'latex'
    MEDIAWIKI = 'mediawiki'
    XLSX = 'xlsx'


class BaseFormatter(abc.ABC):
    _columns: typing.ClassVar = ['Id', 'Song', 'Performed']

    @abc.abstractmethod
    def __init__(self, title: str) -> None: ...

    @abc.abstractmethod
    def add_row(self, row: list[str]) -> None: ...

    @abc.abstractmethod
    def done(self) -> None: ...


class AsciiFormatter(BaseFormatter):
    def __init__(
        self, title: str, *, output_format: str, filename: pathlib.Path | None = None
    ) -> None:
        self._filename = filename
        self._output_format = output_format
        self._title = f'{title}\n' if output_format == 'text' else ''
        self._table = prettytable.PrettyTable()
        self._table.field_names = self._columns
        self._table.align['Id'] = 'r'
        self._table.align['Song'] = 'l'
        self._table.align['Performed'] = 'r'

    def add_row(self, row: list[str]) -> None:
        self._table.add_row(row)

    def done(self) -> None:
        table_text = self._table.get_formatted_string(  # pyright: ignore[reportUnknownMemberType]
            out_format=self._output_format, print_empty=False
        )
        text = f'{self._title}{table_text}\n'
        if self._filename:
            with self._filename.open('w', encoding='utf-8') as fd:
                fd.write(text)
        else:
            sys.stdout.write(text)


class ExcelFormatter(BaseFormatter):
    def __init__(self, title: str, *, filename: pathlib.Path) -> None:
        self._filename = filename
        self._workbook = openpyxl.Workbook()
        for worksheet in self._workbook.worksheets:
            self._workbook.remove(worksheet)
        self._worksheet = self._workbook.create_sheet(title=title)
        self._alignleft = openpyxl.styles.Alignment(horizontal='left')
        self._alignright = openpyxl.styles.Alignment(horizontal='right')
        self.add_row(self._columns)

    def add_row(self, row: list[str]) -> None:
        self._worksheet.append(row)
        self._worksheet.cell(
            row=self._worksheet.max_row, column=1
        ).alignment = self._alignright
        self._worksheet.cell(
            row=self._worksheet.max_row, column=2
        ).alignment = self._alignleft
        self._worksheet.cell(
            row=self._worksheet.max_row, column=3
        ).alignment = self._alignright

    def done(self) -> None:
        for i, column in enumerate(self._worksheet.columns):
            self._worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(i + 1)
            ].width = 1 + max(len(str(cell.value)) for cell in column)
        self._workbook.save(self._filename)


class ChurchToolsSongStatistics:
    def __init__(self, cta: ChurchToolsAPI, config: Configuration) -> None:
        self.cta = cta
        self._log = config.log

    def song_usage(
        self,
        from_date: datetime.datetime,
        to_date: datetime.datetime,
        *,
        output_file: pathlib.Path | None = None,
        output_format: str,
    ) -> None:
        self._log.info('Building song usage statistics')

        year_range = (
            f'{from_date.year}-{to_date.year}'
            if from_date.year != to_date.year
            else f'{from_date.year}'
        )
        title = f'Song statistics for {year_range}'
        if output_format == 'xlsx':
            if not output_file:
                sys.stderr.write(
                    'Error: Format "xlsx" requires to specify an output file\n'
                )
                sys.exit(1)
            formatter = ExcelFormatter(title=title, filename=output_file)
        else:
            formatter = AsciiFormatter(
                title=title, output_format=output_format, filename=output_file
            )

        # Iterate over events and songs and count usage.
        song_counts: dict[tuple[int, str], int] = defaultdict(int)
        with alive_progress.alive_bar(
            title='Calculating statistics', spinner=None, receipt=False
        ) as bar:  # pyright: ignore[reportUnknownVariableType]
            for event in self.cta.get_events(from_date, to_date):
                _, songs = self.cta.get_songs(event, require_tags=False)
                for song in songs:
                    song_counts[song.id, song.name or f'#{song.id}'] += 1
                    bar()

        for (song_id, song_name), count in sorted(
            song_counts.items(), key=lambda s: (-s[1], s[0][1])
        ):
            formatter.add_row([f'#{song_id}', song_name, f'{count}'])

        # Output according to the selected formatter.
        formatter.done()
